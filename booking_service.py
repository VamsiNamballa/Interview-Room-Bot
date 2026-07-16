from __future__ import annotations

from contextlib import contextmanager
from datetime import date, datetime, time, timedelta
from pathlib import Path
from uuid import uuid4

from filelock import FileLock
from openpyxl import Workbook, load_workbook


HEADERS = [
    "booking_id", "candidate_name", "room", "date", "start_time", "end_time",
    "duration_minutes", "implementation_partner", "client", "role", "status",
    "created_at",
]

ROOM_ALIASES = {
    "room 1": "Room 1", "upper room": "Room 1", "top floor": "Room 1",
    "setup room": "Room 1", "room1": "Room 1",
    "room 2": "Room 2", "hall": "Room 2", "room2": "Room 2",
}


class BookingError(ValueError):
    pass


class BookingService:
    def __init__(self, path: str | Path = "data/bookings.xlsx"):
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.lock = FileLock(str(self.path) + ".lock", timeout=10)
        self._ensure_workbook()

    def _ensure_workbook(self) -> None:
        with self.lock:
            if self.path.exists():
                return
            wb = Workbook()
            ws = wb.active
            ws.title = "Bookings"
            ws.append(HEADERS)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:L1"
            widths = [16, 22, 12, 12, 12, 12, 18, 24, 22, 24, 12, 22]
            for index, width in enumerate(widths, start=1):
                ws.column_dimensions[chr(64 + index)].width = width
            wb.save(self.path)

    @staticmethod
    def normalize_room(value: str) -> str:
        cleaned = " ".join(value.lower().strip().split())
        if cleaned not in ROOM_ALIASES:
            raise BookingError("Choose Room 1 (upper room) or Room 2 (hall).")
        return ROOM_ALIASES[cleaned]

    @staticmethod
    def validate_date(day: date) -> None:
        today = date.today()
        if not today <= day <= today + timedelta(days=4):
            raise BookingError("Bookings are allowed only from today through the next four days.")

    @contextmanager
    def _workbook(self, save: bool = False):
        with self.lock:
            wb = load_workbook(self.path)
            try:
                yield wb, wb["Bookings"]
                if save:
                    wb.save(self.path)
            finally:
                wb.close()

    @staticmethod
    def _row_to_dict(headers, row):
        result = dict(zip(headers, row))
        if isinstance(result.get("date"), datetime):
            result["date"] = result["date"].date()
        if isinstance(result.get("start_time"), datetime):
            result["start_time"] = result["start_time"].time()
        if isinstance(result.get("end_time"), datetime):
            result["end_time"] = result["end_time"].time()
        return result

    def list_bookings(self, start_day: date | None = None, active_only: bool = True):
        start_day = start_day or date.today()
        with self._workbook() as (_, ws):
            rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        bookings = [self._row_to_dict(rows[0], row) for row in rows[1:]]
        return sorted(
            [b for b in bookings if b["date"] >= start_day and (not active_only or b["status"] == "confirmed")],
            key=lambda b: (b["date"], b["start_time"], b["room"]),
        )

    def conflicts(self, room: str, day: date, start: time, duration_minutes: int):
        room = self.normalize_room(room)
        requested_start = datetime.combine(day, start)
        requested_end = requested_start + timedelta(minutes=duration_minutes)
        matches = []
        for booking in self.list_bookings(day):
            if booking["date"] != day or booking["room"] != room:
                continue
            existing_start = datetime.combine(day, booking["start_time"])
            existing_end = datetime.combine(day, booking["end_time"])
            if requested_start < existing_end and requested_end > existing_start:
                matches.append(booking)
        return matches

    def create_booking(self, *, candidate_name: str, room: str, day: date, start: time,
                       duration_minutes: int, implementation_partner: str, client: str, role: str):
        self.validate_date(day)
        if not 15 <= duration_minutes <= 240:
            raise BookingError("Duration must be between 15 minutes and 4 hours.")
        required = {
            "candidate name": candidate_name, "implementation partner": implementation_partner,
            "client": client, "role": role,
        }
        missing = [label for label, value in required.items() if not value.strip()]
        if missing:
            raise BookingError("Missing: " + ", ".join(missing) + ".")
        room = self.normalize_room(room)
        end_dt = datetime.combine(day, start) + timedelta(minutes=duration_minutes)
        if end_dt.date() != day:
            raise BookingError("The booking must start and end on the same day.")

        # Lock the conflict check and write together to prevent double-booking.
        with self._workbook(save=True) as (_, ws):
            rows = list(ws.iter_rows(values_only=True))
            for row in rows[1:]:
                booking = self._row_to_dict(rows[0], row)
                if booking["status"] != "confirmed" or booking["date"] != day or booking["room"] != room:
                    continue
                existing_start = datetime.combine(day, booking["start_time"])
                existing_end = datetime.combine(day, booking["end_time"])
                requested_start = datetime.combine(day, start)
                if requested_start < existing_end and end_dt > existing_start:
                    raise BookingError(
                        f"{room} is already booked from {existing_start:%-I:%M %p} to {existing_end:%-I:%M %p}."
                    )
            booking_id = uuid4().hex[:8].upper()
            ws.append([
                booking_id, candidate_name.strip(), room, day, start, end_dt.time(),
                duration_minutes, implementation_partner.strip(), client.strip(), role.strip(),
                "confirmed", datetime.now(),
            ])
        return booking_id, end_dt.time()

    def cancel_booking(self, booking_id: str) -> bool:
        booking_id = booking_id.strip().upper()
        with self._workbook(save=True) as (_, ws):
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row, 1).value).upper() == booking_id and ws.cell(row, 11).value == "confirmed":
                    ws.cell(row, 11).value = "cancelled"
                    return True
        return False

    def alternatives(self, room: str, day: date, start: time, duration_minutes: int):
        room = self.normalize_room(room)
        other = "Room 2" if room == "Room 1" else "Room 1"
        results = []
        if not self.conflicts(other, day, start, duration_minutes):
            results.append(f"{other} is available at {start.strftime('%-I:%M %p')}.")
        probe = datetime.combine(day, start)
        for _ in range(16):
            probe += timedelta(minutes=15)
            if not self.conflicts(room, day, probe.time(), duration_minutes):
                results.append(f"{room}'s next available time is {probe:%-I:%M %p}.")
                break
        return results
